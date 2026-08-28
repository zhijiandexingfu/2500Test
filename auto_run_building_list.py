# -*- coding: utf-8 -*-
"""
auto_run_building_list —— 纯 Python（无浏览器）版全自动判定
====================================================================

与 main.py 的「浏览器混合」流程等价，但**所有接口调用都改走后台 requests**，
不再启动 Playwright / 不依赖浏览器 cookie。区别仅在于：
    - searchAddress：浏览器点击下拉第一项  → 改为后台 search_address() 取返回列表第一项
    - getInstallInfo：浏览器内 fetch         → 改为后台 get_install_info()（脚本测试）

主流程（与需求一致）：
    1. 输入源 = 全網XGSPON升级进度_A1.xlsx 的「8级地址」列
       每行一个地址 → 作为关键词调 searchAddress → 取返回值的第一个地址；
       若输入楼宇名与系统返回地址不能完全匹配 → remark 写「eshop地址不匹配」。
    2. 调 getAddressDetail → 返回 busiResp 保存为 busiRespObj
       → 取出该楼宇所有 floor / flat。
    3. 采样：楼层先按数字（int）排序、同层 Flat 字典序兜底，再每 5 层随机选 1 层
       curFloor，当层随机选 1 个 curFlat。
    4. 对每组的 (floor, flat) 调 getInstallInfo 自动判定：
         isXGSPON='Y' → Is2500Support=Y,  remark=可下2500M
         isXGSPON='N' → Is2500Support=N,  defeatBuilding=Y, remark=不可下2500M
         其他兜底       → Is2500Support=E,  remark=待排查
    5. 落库 Excel：输出表头 = A1 原表头 + 追加字段（curBuilding/curFloor/curFlat/
       Is2500Support/defeatBuilding/remark/buildingCode/ofcaCode/carrier …），
       行键幂等 + 楼宇级/行级断点续跑（复用 storage 模块）。

用法：
    python auto_run_building_list.py
    python auto_run_building_list.py --input "全網XGSPON升级进度_A1.xlsx" --limit 5
    python auto_run_building_list.py --redo
    python auto_run_building_list.py --token-search <t> --token-detail <t> --token-install <t>
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import config
import openpyxl
from api_client import ApiError, CmhkApiClient
from storage import ExcelStorage, ProgressTracker

# 复用 main.py 里已验证的纯函数（不依赖浏览器）
from main import (
    build_install_bus_info,
    extract_building_name,
    is_address_matched,
    pick_building_code,
    pick_sampled_units,
    summarize_install,
)


logger = logging.getLogger("xgps.auto_run_building_list")

# 默认输入：全網XGSPON升级进度_A1.xlsx（与脚本同目录的上一级 Downloads）
DEFAULT_INPUT = r"C:\Users\zengh\Downloads\全網XGSPON升级进度_A1.xlsx"
DEFAULT_COLUMN = "8级地址"
DEFAULT_OUTPUT = config.EXCEL_PATH.replace("result.xlsx", "auto_run_building_list_result.xlsx")

# 追加到 A1 原表头之后的字段（输出格式由需求规定）
APPENDED_HEADERS = [
    "inputBuilding",   # 8级地址的值，用作行键（ExcelStorage 固定按此列定位）
    "curBuilding",     # 页面选中的地址文本
    "curFloor",        # 采样楼层
    "curFlat",         # 采样单位
    "Is2500Support",   # Y / N / E
    "defeatBuilding",  # 不可卖时为 Y
    "remark",          # 判定说明 / 错误信息
    "buildingCode",    # 楼宇编码
    "ofcaCode",        # OFCA 编码
    "carrier",         # 覆盖运营商（getInstallInfo 返回）
    "isXGSPONsupport", # 接口侧 XGS-PON(2500M) 支持标识
    "coverType",       # 覆盖类型（如 GP）
    "floorGroup",      # 采样分组（每 5 层一组）
    "level11_address", # 11级地址 = curBuilding,curFloor层curFlat户
    "checkedAt",       # 本行采集时间
]


# ----------------------------------------------------------------------
# 输入读取：从 A1 Excel 提取「8级地址」列整行（去重，保留顺序）
# ----------------------------------------------------------------------
def read_address_rows(
    path: str,
    column: str = DEFAULT_COLUMN,
    sheet: Optional[str] = None,
) -> Tuple[List[str], List[Tuple[str, Dict[str, Any]]]]:
    """
    读取 Excel 指定工作表的「8级地址」列，返回 (A1表头列表, [(关键词, 整行dict), ...])。

    整行 dict 保留 A1 的全部原始列（區域/優先級/Site/.../下挂楼宇/8级地址），
    以便输出时原样携带。关键词去重、保留首次出现顺序。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    headers = [c.value for c in ws[1]]
    if column not in headers:
        raise ValueError(
            f"输入文件缺少「{column}」列；现有表头：{headers}"
        )
    col = headers.index(column) + 1

    seen = set()
    rows: List[Tuple[str, Dict[str, Any]]] = []
    dup = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        kw = str(v).strip()
        if not kw:
            continue
        if kw in seen:
            dup += 1
            continue
        seen.add(kw)
        row = {
            headers[c]: (ws.cell(r, c + 1).value or "")
            for c in range(len(headers))
        }
        rows.append((kw, row))
    wb.close()
    logger.info("输入提取：唯一地址 %d 个，去重 %d 条重复行", len(rows), dup)
    return headers, rows


def _read_header_row(path: str) -> List[str]:
    """读取已有结果文件的首行表头（用于 schema 一致性校验）。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    wb.close()
    return headers


def base_record(keyword: str, input_row: Dict[str, Any]) -> Dict[str, Any]:
    """以 A1 整行为底，附上行键列 inputBuilding（=8级地址值）。"""
    rec = {h: (input_row.get(h, "") or "") for h in input_row}
    rec["inputBuilding"] = keyword
    return rec


def apply_auto_result_script(record: Dict[str, Any]) -> Dict[str, Any]:
    """按接口 isXGSPONsupport 自动填充判定列（脚本版固定文案）。"""
    flag = str(record.get("isXGSPONsupport", "") or "").strip()
    if flag == "Y":
        record.update({"Is2500Support": "Y", "remark": "可下2500M"})
    elif flag == "N":
        record.update({"Is2500Support": "N", "defeatBuilding": "Y", "remark": "不可下2500M"})
    else:
        record.update({"Is2500Support": "E", "remark": "待排查"})
    return record


# ----------------------------------------------------------------------
# 单栋楼宇处理（纯 API，无浏览器）
# ----------------------------------------------------------------------
def process_building(
    client: CmhkApiClient,
    storage: ExcelStorage,
    keyword: str,
    input_row: Dict[str, Any],
) -> bool:
    """
    处理一栋楼宇：searchAddress → getAddressDetail → 采样 → getInstallInfo → 落库。
    :return: True=全部样本成功；False=存在失败（下次运行会重试该楼宇）
    """
    base = base_record(keyword, input_row)

    # ---- 1. searchAddress：取返回列表第一项 ----
    try:
        data = client.search_address(keyword)
        items = (data.get("busiResp") or {}).get("busiDataResp") or []
    except ApiError as exc:
        msg = f"searchAddress: {exc}"
        logger.warning(msg)
        storage.upsert_row({**base, "remark": f"错误:{msg[:120]}"})
        return False
    if not items:
        # 系统无该地址候选：按「eshop地址不匹配」降级
        print(f"    [降级] eshop地址不匹配（keyword={keyword}）")
        storage.upsert_row({**base, "curBuilding": "", "remark": "eshop地址不匹配"})
        return True

    address_obj = items[0]
    cur_building = address_obj.get("value", "")
    print(f"    curBuilding = {cur_building}")

    # ---- 1.5 地址匹配校验：系统返回地址须包含输入楼宇名 ----
    if not is_address_matched(keyword, cur_building):
        name = extract_building_name(keyword)
        print(f"    [降级] eshop地址不匹配（输入[{name}]，系统最近似返回[{cur_building}]）")
        logger.warning("地址未匹配(eshop地址不匹配): %s -> %s", keyword, cur_building)
        storage.upsert_row({**base, "curBuilding": cur_building, "remark": "eshop地址不匹配"})
        return True

    client.throttle()

    # ---- 2. getAddressDetail → busiRespObj ----
    try:
        detail_data = client.get_address_detail(address_obj)
    except ApiError as exc:
        msg = f"getAddressDetail: {exc}"
        print(f"    [接口异常] {msg}")
        storage.upsert_row({**base, "curBuilding": cur_building, "remark": f"错误:{msg[:120]}"})
        return False
    busi_resp_obj = detail_data.get("busiResp") or {}

    # ---- 3. 采样（数字优先排序 + 每 5 层随机 1 单位）----
    samples = pick_sampled_units(busi_resp_obj, seed=keyword)
    if not samples:
        print("    [提示] 该楼宇无楼层/单位数据")
        storage.upsert_row({**base, "curBuilding": cur_building, "remark": "楼宇无楼层单位数据"})
        return True

    print(f"    共采样 {len(samples)} 组: "
          + ", ".join(f"{f}层{p}室(组{g})" for f, p, g in samples))

    building_code = pick_building_code(address_obj.get("carrierInfo", []))
    all_ok = True

    # ---- 4+5. 每组：getInstallInfo（纯 API） → 自动判定 → 落库 ----
    for idx, (floor, flat, group_label) in enumerate(samples, start=1):
        if storage.has_row(keyword, floor, flat):
            print(f"    [{idx}/{len(samples)}] {floor}层{flat}室 已存在，跳过")
            continue

        record: Dict[str, Any] = {
            **base,
            "curBuilding": cur_building,
            "curFloor": floor,
            "curFlat": flat,
            "buildingCode": building_code,
            "ofcaCode": address_obj.get("ofcaCode", ""),
            "floorGroup": group_label,
        }

        bus_info = build_install_bus_info(address_obj, building_code, floor, flat)
        try:
            install_data = client.get_install_info(bus_info)
        except ApiError as exc:
            print(f"    [{idx}/{len(samples)}] {floor}层{flat}室 "
                  f"[getInstallInfo 失败] {exc}")
            record.update({"remark": f"错误:getInstallInfo:{str(exc)[:100]}"})
            storage.upsert_row(record)
            all_ok = False
            client.throttle()
            continue

        record.update(summarize_install(install_data))
        record = apply_auto_result_script(record)
        print(f"    [{idx}/{len(samples)}] {floor}层{flat}室(组{group_label}) "
              f"isXGSPONsupport={record['isXGSPONsupport'] or '-'} "
              f"-> {record['remark']}")
        storage.upsert_row(record)
        client.throttle()

    return all_ok


# ----------------------------------------------------------------------
# 主流程
# ----------------------------------------------------------------------
def run(args: argparse.Namespace) -> int:
    try:
        a1_headers, rows = read_address_rows(args.input, column=args.column, sheet=args.sheet)
    except Exception as exc:
        print(f"[错误] 读取输入失败：{exc}")
        return 1
    if not rows:
        print(f"[错误] 输入文件「{args.input}」的「{args.column}」列无有效数据")
        return 1
    if args.limit:
        rows = rows[: args.limit]
    total = len(rows)

    # 输出表头 = A1 原表头 + 追加字段
    output_headers = a1_headers + APPENDED_HEADERS

    # schema 一致性：若已有结果文件表头与本次不一致，--redo 重建，否则报错退出
    if os.path.exists(args.output):
        existing = _read_header_row(args.output)
        if existing != output_headers:
            if args.redo:
                os.remove(args.output)
                print("[--redo] 检测到表头变化，已重建结果文件")
            else:
                print("[错误] 结果文件表头与本次输入不一致；"
                      "请换 --output 路径或加 --redo 重建")
                return 1

    storage = ExcelStorage(path=args.output, headers=output_headers)
    progress = ProgressTracker(ignore_existing=args.redo)

    client = CmhkApiClient(
        search_token=args.token_search,
        detail_token=args.token_detail,
        install_token=args.token_install,
    )

    print(
        f"[启动] 输入 {total} 栋 | Excel: {args.output} | "
        f"已完成 {progress.completed_count()} 栋 | "
        f"模式: 纯Python全自动（无浏览器） | "
        f"采样规则: 每 {config.FLOOR_GROUP_SIZE} 层随机 1 层 + 当层随机 1 单位"
    )

    failed: List[str] = []
    for row_no, (keyword, input_row) in enumerate(rows, start=1):
        if not args.redo and progress.is_done(keyword):
            print(f"[跳过] ({row_no}/{total}) {keyword} 已完成")
            continue

        print(f"\n[{row_no}/{total}] {keyword}")
        try:
            ok = process_building(client, storage, keyword, input_row)
        except KeyboardInterrupt:
            print("\n[中断] 用户终止，已落库数据保留")
            raise
        except Exception as exc:
            logger.exception("处理楼宇异常: %s", keyword)
            storage.upsert_row({**base_record(keyword, input_row),
                                "remark": f"错误:未预期异常:{str(exc)[:100]}"})
            ok = False

        if ok:
            progress.mark_done(keyword)
        else:
            failed.append(keyword)
        time.sleep(config.REQUEST_INTERVAL)

    print(f"\n[完成] 成功 {total - len(failed)}/{total}"
          + (f"，失败 {len(failed)} 栋" if failed else ""))
    if failed:
        print("失败楼宇：")
        for f in failed:
            print(f"  - {f}")
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="纯Python版全自动判定：读A1表8级地址→searchAddress→getAddressDetail→采样→getInstallInfo→落库",
    )
    p.add_argument("--input", default=DEFAULT_INPUT, help="输入 Excel（含「8级地址」列）")
    p.add_argument("--sheet", default=None, help="工作表名（默认取活动表）")
    p.add_argument("--column", default=DEFAULT_COLUMN, help="地址列名，默认「8级地址」")
    p.add_argument("--output", default=DEFAULT_OUTPUT, help="结果 Excel 路径")
    p.add_argument("--limit", type=int, default=0, help="只跑前 N 栋（调试用）")
    p.add_argument("--redo", action="store_true", help="忽略已有进度，全部重跑")
    p.add_argument("--token-search", default=config.XGIOG_TOKEN_SEARCH)
    p.add_argument("--token-detail", default=config.XGIOG_TOKEN_DETAIL)
    p.add_argument("--token-install", default=config.XGIOG_TOKEN_INSTALL)
    return p


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    sys.exit(run(build_arg_parser().parse_args()))
