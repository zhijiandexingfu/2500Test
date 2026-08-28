# -*- coding: utf-8 -*-
"""
结果落库模块（Excel 追加写入 + 断点续跑）
==========================================

数据模型（与旧版「一行输入对一行输出」不同）：
    * 一栋楼（inputBuilding）经「每 5 层随机选 1 层」采样后会产生【多行】结果，
      因此 Excel 采用「按行键追加」模型：
          行键 rowKey = (inputBuilding, curFloor, curFlat)
    * upsert_row()：键已存在则原位更新，不存在则追加到末尾 —— 幂等，重跑不产生重复行。
    * 断点续跑分两层：
          1. 楼宇级：progress.json 记录已完整跑完的 inputBuilding，重启后整栋跳过；
          2. 行级：  程序在楼宇中途被杀时，已写入的采样行不丢（键存在即跳过），
                     该楼宇因未写 progress 标记会在下次运行时补齐缺失的采样行。
    * 实时刷新：每写一行立即 save()，随时可用 Excel/WPS 打开查看最新进度。

可扩展：列结构由 config.BASE_HEADERS + EXTRA_HEADERS 决定，增删列无需改本模块。
"""

import json
import os
import threading
from datetime import datetime

from openpyxl import Workbook, load_workbook

import config


class ExcelStorage:
    """排查结果 Excel 落库（线程安全 + 行级幂等）"""

    def __init__(self, path: str = None, headers: list = None):
        self.path = path or config.EXCEL_PATH
        self.headers = headers or (config.BASE_HEADERS + config.EXTRA_HEADERS)
        self._lock = threading.Lock()                 # 防止多线程写坏文件
        self._col_index = {h: i + 1 for i, h in enumerate(self.headers)}
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        self._load_or_create()
        self._existing_keys = self._scan_keys()       # {(inputBuilding, floor, flat): 行号}

    # ------------------------------------------------------------------
    # 初始化 / 加载
    # ------------------------------------------------------------------
    def _load_or_create(self) -> None:
        """加载已有 Excel；不存在则新建并写表头"""
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
            # 兼容旧文件缺列的情况：补齐表头
            for i, h in enumerate(self.headers, start=1):
                if self.ws.cell(row=1, column=i).value != h:
                    self.ws.cell(row=1, column=i, value=h)
            self.wb.save(self.path)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "排查结果"
            for i, h in enumerate(self.headers, start=1):
                self.ws.cell(row=1, column=i, value=h)
            self.wb.save(self.path)

    def _scan_keys(self) -> dict:
        """扫描已有数据行，建立 行键 -> Excel物理行号 索引（断点续跑依据）"""
        keys = {}
        for row in range(2, self.ws.max_row + 1):
            input_b = self.ws.cell(row=row, column=self._col_index["inputBuilding"]).value
            if input_b is None:
                continue
            floor = self.ws.cell(row=row, column=self._col_index["curFloor"]).value or ""
            flat = self.ws.cell(row=row, column=self._col_index["curFlat"]).value or ""
            keys[(str(input_b), str(floor), str(flat))] = row
        return keys

    # ------------------------------------------------------------------
    # 写入（幂等：行键存在则更新，否则追加）
    # ------------------------------------------------------------------
    def upsert_row(self, record: dict) -> int:
        """
        写入（或覆盖）一行采样结果并立即保存。

        :param record: {列名: 值} 字典，缺省列留空
        :return: 写入的 Excel 物理行号
        """
        with self._lock:
            key = (str(record.get("inputBuilding", "")),
                   str(record.get("curFloor", "")),
                   str(record.get("curFlat", "")))
            excel_row = self._existing_keys.get(key, self.ws.max_row + 1)
            # 先清空旧行，保证重跑时无残留脏数据
            for col in range(1, len(self.headers) + 1):
                self.ws.cell(row=excel_row, column=col, value=None)
            # 自动派生「11级地址」= curBuilding,curFloor层curFlat户
            # 只要有 curFloor 或 curFlat 就拼，避免空记录里冒出裸字串
            cb = record.get("curBuilding", "") or ""
            cf = record.get("curFloor", "")
            ct = record.get("curFlat", "")
            if cf or ct:
                record.setdefault(
                    "level11_address",
                    f"{cb},{cf}层{ct}户".lstrip(","),
                )
            for name, value in record.items():
                col = self._col_index.get(name)
                if col:
                    self.ws.cell(row=excel_row, column=col, value=value)
            # 自动补充采集时间（record 未显式给 checkedAt 时）
            if not record.get("checkedAt"):
                self.ws.cell(
                    row=excel_row, column=self._col_index["checkedAt"],
                    value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                )
            self.wb.save(self.path)   # 实时落盘
            self._existing_keys[key] = excel_row
            return excel_row

    def has_row(self, input_building: str, floor: str, flat: str) -> bool:
        """行键是否已存在（用于行级断点续跑跳过）"""
        return (str(input_building), str(floor), str(flat)) in self._existing_keys

    def row_count(self) -> int:
        """当前数据行数（不含表头）"""
        return len(self._existing_keys)


# ----------------------------------------------------------------------
# 楼宇级断点进度（progress.json）
# ----------------------------------------------------------------------
class ProgressTracker:
    """记录已完整跑完的楼宇名称集合，实现楼宇级断点续跑"""

    def __init__(self, path: str = None, ignore_existing: bool = False):
        """
        :param ignore_existing: True 时不加载已有进度文件（配合 --redo 全部重跑），
                                完成标记会写入临时文件，不污染原进度。
        """
        self.path = path or config.PROGRESS_PATH
        if ignore_existing:
            base, ext = os.path.splitext(self.path)
            self.path = f"{base}.redo{ext}"
        self._lock = threading.Lock()
        self._done: set = set()
        if not ignore_existing and os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    self._done = set(json.load(f).get("done", []))
            except Exception:
                self._done = set()   # 进度文件损坏时不阻塞，重跑已完成的楼宇

    def is_done(self, input_building: str) -> bool:
        return input_building in self._done

    def mark_done(self, input_building: str) -> None:
        """标记楼宇完成并立即落盘"""
        with self._lock:
            self._done.add(input_building)
            os.makedirs(os.path.dirname(self.path), exist_ok=True)
            with open(self.path, "w", encoding="utf-8") as f:
                json.dump({"done": sorted(self._done)}, f, ensure_ascii=False, indent=1)

    def completed_count(self) -> int:
        return len(self._done)
